'''
    以数据表格的形式，将整体调度方案进行可视化展示，包括各晶圆的加工路径、每一晶圆每一道工序任务各阶段的时间安排、
每一加工腔执行所分配任务的时间安排、机械臂进行调度时的操作时刻和操作目的，分别进行了详尽的展示。
调用里面的方法，传入分配方案、实时调度模块的结果，即可方便的导出各项数据。
'''
# 使用openpyxl生成xlsx的excel文件
from Solution_Algorithm.data_source import *
import xlwt
import copy
import pandas as pd


class Excel:
    # 利用python的xlwt模块自适应列宽写入exceL
    @staticmethod
    def reset_col(filename):
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        wb = load_workbook(filename)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
        df = pd.read_excel(filename, sheet).fillna('-')
        df.loc[len(df)] = list(df.columns)  # 把标题行附加到最后一行
        for col in df.columns:
            index = list(df.columns).index(col)  # 列序号
            letter = get_column_letter(index + 1)  # 列字母
            collen = df[col].apply(lambda x: len(str(x).encode())).max()  # 获取这一列长度的最大值 当然也可以用min获取最小值 mean获取平均值
            ws.column_dimensions[letter].width = collen + 2  # 也就是列宽为最大长度*1.2 可以自己调整
        wb.save(filename)

    @staticmethod
    # 获取每列所占用的最大列宽
    def get_max_col(max_list):
        line_list = []
        # i表示行，j代表列
        for j in range(len(max_list)):
            line_list.append(max(max_list[j]))  # 将每列最大宽度存入line_list
        return line_list

    @staticmethod
    def write_excel_y(data, filename, et):  # 将列表一列一列展示
        row_num = 0  # 记录写入行数
        col_list = []  # 记录每行宽度
        # 创建一个Workbook对象
        book = xlwt.Workbook(encoding="utf-8", style_compression=0)
        # 创建一个sheet对象
        sheet = book.add_sheet("完工时间：{}s".format(et), cell_overwrite_ok=True)
        # 写入数据
        for i in range(0, len(data)):
            col_num = []
            for j in range(0, len(data[i])):
                sheet.write(j, i, data[i][j])
                col_num.append(len(data[i][j].encode('gb18030')))  # 计算每列值的大小
            col_list.append(copy.copy(col_num))  # 记录一行每列写入的长度
        # 获取每列最大宽度
        col_max_num = Excel.get_max_col(col_list)
        # 设置自适应列宽
        for i in range(0, len(col_max_num)):
            # 256*字符数得到excel列宽,为了不显得特别紧凑添加两个字符宽度
            sheet.col(i).width = 256 * (col_max_num[i] + 2)
        # 保存excel文件
        book.save("{}.xlsx".format(filename))

    @staticmethod
    def write_excel_x(data, filename, et):  # 将列表一行一行展示
        row_num = 0  # 记录写入行数
        col_list = []  # 记录每行宽度
        # 创建一个Workbook对象
        book = xlwt.Workbook(encoding="utf-8", style_compression=0)
        # 创建一个sheet对象
        sheet = book.add_sheet("完工时间：{}s".format(et), cell_overwrite_ok=True)
        # 写入数据
        for i in range(0, len(data)):
            col_num = []
            for j in range(0, len(data[i])):
                sheet.write(i, j, data[i][j])
                col_num.append(len(data[i][j].encode('gb18030')))  # 计算每列值的大小
            col_list.append(copy.copy(col_num))  # 记录一行每列写入的长度
        # 获取每列最大宽度
        col_max_num = Excel.get_max_col(col_list)
        # 设置自适应列宽
        for i in range(0, len(col_max_num)):
            # 256*字符数得到excel列宽,为了不显得特别紧凑添加两个字符宽度
            sheet.col(i).width = 256 * (col_max_num[i] + 2)
        # 保存excel文件
        book.save("{}.xlsx".format(filename))

    @staticmethod
    def out_order(O, lujing, et):
        from openpyxl import Workbook
        top = ["完工时间：{}s".format(et)]
        columns = ['工序编号', '加工腔', '到达时间', '卸载晶圆', '加工时间', '驻留时间', '取走晶圆', '结束', '到达下一工序经历时间']
        datas = []
        for i in range(1, len(O), 1):
            for j in range(1, len(O[i]), 1):
                datas.append(["({},{})".format(i, j), lujing[i - 1][j - 1], str(O[i][j].begin), str(O[i][j].xie),
                              str(O[i][j].process),
                              str(O[i][j].nei_wait), str(O[i][j].qu), str(O[i][j].end),
                              str(O[i][j + 1].wai_wait) if j != M else 'X'])
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = ("{}工艺路径下晶圆各工序起止时间".format(Tp_strname()))
        sheet.append(top)
        sheet.append(columns)
        for data in datas:
            sheet.append(data)
        workbook.save("./Result/{}工艺路径/{}工艺路径晶圆各工序起止时间({}).xlsx".format(Tp_strname(), Tp_strname(), N))

    @staticmethod
    def out_PM(O, lujing, et):
        # 将工艺路径用到的腔室转为字典，用来记录腔室加工过的晶圆编号
        machines = Machines(Tp)
        arr = []
        for i in range(len(machines)):
            for j in range(len(machines[i])):
                arr.append(machines[i][j])
        PM_wafers = {x: ["{}".format(x)] for x in arr}
        PM_wafers1 = {x: [] for x in arr}  # 记录加工过的晶圆编号
        for i in range(1, len(O), 1):
            for j in range(1, len(O[i]), 1):
                Ms = list(lujing[i - 1])[j - 1]
                # 记录此腔室加工过的晶圆编号
                PM_wafers1[Ms].append([i, j])
        for pm in PM_wafers1:
            a = 0.0
            for m in PM_wafers1[pm]:
                PM_wafers[pm].append("{}到{}s:空闲".format(a, O[m[0]][m[1]].begin))
                PM_wafers[pm].append("{}到{}s:处理晶圆{}".format(O[m[0]][m[1]].begin, O[m[0]][m[1]].end, m[0]))
                a = O[m[0]][m[1]].end
            PM_wafers[pm].append("{}s:完工".format(a))
        datas = [PM_wafers[pm] for pm in PM_wafers]
        Excel.write_excel_y(datas, "./Result/{}工艺路径/{}工艺路径加工腔调度方案({})".format(Tp_strname(), Tp_strname(), N), et)

    @staticmethod
    def out_TM(TM_caozuo, et):
        from openpyxl import Workbook
        top = ["完工时间：{}s".format(et)]
        columns = ['TM1操作时刻', 'TM1操作', 'TM2操作时刻', 'TM2操作', 'TM3操作时刻', 'TM3操作']
        datas = []
        for i in range(max([len(TM_caozuo[0]), len(TM_caozuo[1]), len(TM_caozuo[2])])):
            datas.append([' ', ' ', ' ', ' ', ' ', ' '])
        for j in range(len(TM_caozuo[0])):
            datas[j][0] = "{}至{}s".format(TM_caozuo[0][j].begin, TM_caozuo[0][j].end)
            datas[j][1] = TM_caozuo[0][j].function
        for j in range(len(TM_caozuo[1])):
            datas[j][2] = "{}至{}s".format(TM_caozuo[1][j].begin, TM_caozuo[1][j].end)
            datas[j][3] = TM_caozuo[1][j].function
        for j in range(len(TM_caozuo[2])):
            datas[j][4] = "{}至{}s".format(TM_caozuo[2][j].begin, TM_caozuo[2][j].end)
            datas[j][5] = TM_caozuo[2][j].function
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = ("{}工艺路径下机械臂调度方案".format(Tp_strname()))
        sheet.append(top)
        sheet.append(columns)
        for data in datas:
            sheet.append(data)
        workbook.save("./Result/{}工艺路径/{}工艺路径机械臂调度方案({}).xlsx".format(Tp_strname(), Tp_strname(), N))
        # Excel.write_excel_x(datas, "{}工艺路径机械臂调度方案".format(Tp_strname()))
        Excel.reset_col("./Result/{}工艺路径/{}工艺路径机械臂调度方案({}).xlsx".format(Tp_strname(), Tp_strname(), N))

    @staticmethod
    def out_lujing(lujing, et):
        from openpyxl import Workbook
        top = ["完工时间：{}s".format(et)]
        columns = ['晶圆编号'] + ["第{}道工序".format(i) for i in range(1, len(lujing[0]) + 1)]
        datas = []
        for j in range(0, len(lujing), 1):
            datas.append(["晶圆{}".format(j + 1)] + lujing[j])
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = ("{}工艺路径下各晶圆加工路径".format(Tp_strname()))
        sheet.append(top)
        sheet.append(columns)
        for data in datas:
            sheet.append(data)
        workbook.save("./Result/{}工艺路径/{}工艺路径各晶圆加工路径({}).xlsx".format(Tp_strname(), Tp_strname(), N))

    @staticmethod
    def out_clear(clear, et):
        from openpyxl import Workbook
        top = ["完工时间：{}s".format(et)]
        datas = [([pm] + clear[pm]) for pm in clear]
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = ("{}工艺路径加工腔清洗起止时刻".format(Tp_strname()))
        sheet.append(top)
        for data in datas:
            sheet.append(data)
        workbook.save("./Result/{}工艺路径/{}工艺路径加工腔清洁起止时刻({}).xlsx".format(Tp_strname(), Tp_strname(), N))
        Excel.reset_col("./Result/{}工艺路径/{}工艺路径加工腔清洁起止时刻({}).xlsx".format(Tp_strname(), Tp_strname(), N))
